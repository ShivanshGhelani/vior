import cv2
from ultralytics import YOLO
import pandas as pd
import os
from glob import glob
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import time

def print_menu():
    """Display the main menu options"""
    print("\n=== Object Detection Analysis System ===")
    print("1. Process Images (Generate Excel report)")
    print("2. Process Videos (Generate Excel report)")
    print("3. Process Both Images and Videos")
    print("4. Exit")
    return input("\nSelect an option (1-4): ")

def get_object_positions(image_path, model_path="models/yolov8l.pt"):
    """
    Process an image and return dictionary of detected objects with their positions
    Args:
        image_path: Path to the image file
        model_path: Path to the YOLO model file
    Returns:
        List of detections with their details including object IDs
    """
    # Load model
    model = YOLO(model_path)
    
    # Read image
    frame = cv2.imread(image_path)
    if frame is None:
        raise ValueError(f"Could not read image at {image_path}")
    
    # Get image dimensions
    height, width = frame.shape[:2]
    
    # Run detection
    results = model(frame)
    
    # Dictionary to keep track of object counts
    object_counts = {}
    
    # List to store all detections
    detections = []
    
    for result in results:
        for box in result.boxes:
            x1, y1, x2, y2 = box.xyxy[0].tolist()
            cls = int(box.cls[0].item())
            label = model.names[cls]
            confidence = float(box.conf[0].item())
            
            # Update object count and get unique ID
            object_counts[label] = object_counts.get(label, 0) + 1
            object_id = f"{label}_{object_counts[label]}"
            
            # Calculate center points
            x_center = (x1 + x2) / 2
            y_center = (y1 + y2) / 2
            
            # Determine horizontal position
            if x_center < width * 0.3:
                h_pos = "left"
            elif x_center > width * 0.7:
                h_pos = "right"
            else:
                h_pos = "centre"
            
            # Determine vertical position
            if y_center < height * 0.3:
                v_pos = "top"
            elif y_center > height * 0.7:
                v_pos = "bottom"
            else:
                v_pos = "centre"
            
            # Combine positions for final position string
            if v_pos == "centre" and h_pos == "centre":
                position = "centre"
            elif v_pos == "centre":
                position = f"centre-{h_pos}"
            elif h_pos == "centre":
                position = v_pos
            else:
                position = f"{v_pos}-{h_pos}"
            
            # Add detection to list with object ID
            detections.append({
                'image_name': os.path.basename(image_path),
                'object_id': object_id,
                'object': label,
                'position': position,
                'confidence': round(confidence, 3),
                'bbox_x1': round(x1, 2),
                'bbox_y1': round(y1, 2),
                'bbox_x2': round(x2, 2),
                'bbox_y2': round(y2, 2)
            })
    
    return detections

def process_video(video_path, model_path="models/yolov8l.pt", sample_rate=1):
    """
    Process a video and track unique object instances with their positions
    """
    # Load model
    model = YOLO(model_path)
    
    # Open video
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise ValueError(f"Could not open video at {video_path}")
    
    # Get video properties
    fps = int(cap.get(cv2.CAP_PROP_FPS))
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    # Dictionary to track unique objects
    unique_objects = {}
    object_counts = {}
    
    print(f"\nProcessing video: {os.path.basename(video_path)}")
    print(f"Video properties: {width}x{height} @ {fps}fps, {frame_count} frames")
    
    frame_number = 0
    
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break
            
        # Process every nth frame
        if frame_number % sample_rate == 0:
            # Run detection
            results = model(frame)
            
            # Process detections
            for result in results:
                for box in result.boxes:
                    x1, y1, x2, y2 = box.xyxy[0].tolist()
                    cls = int(box.cls[0].item())
                    label = model.names[cls]
                    confidence = float(box.conf[0].item())
                    
                    # Calculate center points
                    x_center = (x1 + x2) / 2
                    y_center = (y1 + y2) / 2
                    
                    # Determine horizontal position
                    if x_center < width * 0.3:
                        h_pos = "left"
                    elif x_center > width * 0.7:
                        h_pos = "right"
                    else:
                        h_pos = "centre"
                    
                    # Determine vertical position
                    if y_center < height * 0.3:
                        v_pos = "top"
                    elif y_center > height * 0.7:
                        v_pos = "bottom"
                    else:
                        v_pos = "centre"
                    
                    # Combine positions
                    if v_pos == "centre" and h_pos == "centre":
                        position = "centre"
                    elif v_pos == "centre":
                        position = f"centre-{h_pos}"
                    elif h_pos == "centre":
                        position = v_pos
                    else:
                        position = f"{v_pos}-{h_pos}"
                    
                    # Check if this is a new unique object based on position proximity
                    new_object = True
                    for obj_key, obj_data in unique_objects.items():
                        if obj_data['object'] == label:
                            stored_x = obj_data['center_x']
                            stored_y = obj_data['center_y']
                            # If within proximity of existing object, consider it the same instance
                            if abs(stored_x - x_center) < width * 0.1 and abs(stored_y - y_center) < height * 0.1:
                                new_object = False
                                # Update position if confidence is higher
                                if confidence > obj_data['confidence']:
                                    obj_data['position'] = position
                                    obj_data['confidence'] = confidence
                                    obj_data['center_x'] = x_center
                                    obj_data['center_y'] = y_center
                                break
                    
                    # If it's a new unique object, create new entry
                    if new_object:
                        object_counts[label] = object_counts.get(label, 0) + 1
                        object_id = f"{label}_{object_counts[label]}"
                        unique_objects[object_id] = {
                            'object': label,
                            'object_id': object_id,
                            'position': position,
                            'confidence': confidence,
                            'center_x': x_center,
                            'center_y': y_center
                        }
            
            # Print progress
            if frame_number % (sample_rate * 30) == 0:
                progress = (frame_number / frame_count) * 100
                print(f"Progress: {progress:.1f}%")
        
        frame_number += 1
    
    cap.release()
    
    # Convert tracking data to list format
    tracked_objects = []
    for obj_id, data in unique_objects.items():
        tracked_objects.append({
            'object_id': obj_id,
            'object': data['object'],
            'position': data['position'],
            'confidence': round(data['confidence'], 3)
        })
    
    return tracked_objects

def save_to_excel(detections, output_path='object_detections.xlsx', video_mode=False):
    """
    Save detections to Excel with proper formatting
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if video_mode:
            # Create DataFrame for video tracking data
            df = pd.DataFrame(detections)
            
            # Group objects by type
            object_groups = df.groupby('object')
            
            current_row = 0
            
            for obj_type, group in object_groups:
                # Write object type header
                header_df = pd.DataFrame([[f'Object Type: {obj_type}']], columns=[''])
                header_df.to_excel(writer, startrow=current_row, index=False, header=False)
                
                # Write object instances
                group = group.sort_values('object_id')
                group = group[['object_id', 'position', 'confidence']]
                group.to_excel(writer, startrow=current_row + 1, index=False)
                
                current_row += len(group) + 3
        else:
            # Original image processing logic
            current_row = 0
            for img_group in pd.DataFrame(detections).groupby('image_name'):
                image_name, img_data = img_group
                
                # Write image header
                header_df = pd.DataFrame([['Image: ' + image_name]], columns=[''])
                header_df.to_excel(writer, startrow=current_row, index=False, header=False)
                
                # Group objects by type within each image
                object_groups = img_data.groupby('object')
                current_row += 2
                
                for obj_type, group in object_groups:
                    # Write object type subheader
                    subheader_df = pd.DataFrame([[f'Object Type: {obj_type}']], columns=[''])
                    subheader_df.to_excel(writer, startrow=current_row, index=False, header=False)
                    
                    # Write object instances
                    group = group.sort_values('object_id')
                    group = group[['object_id', 'position', 'confidence']]
                    group.to_excel(writer, startrow=current_row + 1, index=False)
                    
                    current_row += len(group) + 3
                
                current_row += 2
        
        # Apply formatting
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Format headers and adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                    
                    # Format headers
                    if isinstance(cell.value, str) and ('Image:' in cell.value or 'Object Type:' in cell.value):
                        cell.font = Font(bold=True, size=12)
                        cell.fill = PatternFill(start_color='E0E0E0',
                                              end_color='E0E0E0',
                                              fill_type='solid')
            
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

def process_all_images():
    """Process all images in the images folder"""
    image_paths = glob('images/*.jpg') + glob('images/*.png')
    image_paths.sort()
    
    all_image_detections = []
    if not image_paths:
        print("No images found in the 'images' folder!")
        return False
        
    print("\nProcessing images...")
    for img_path in image_paths:
        print(f"\nProcessing {img_path}:")
        try:
            detections = get_object_positions(img_path)
            if detections:
                all_image_detections.extend(detections)
                print(f"Found {len(detections)} objects")
        except Exception as e:
            print(f"Error processing {img_path}: {str(e)}")
    
    if all_image_detections:
        save_to_excel(all_image_detections, 'image_detections.xlsx', video_mode=False)
        print("\nImage results saved to image_detections.xlsx")
        return True
    return False

def process_all_videos():
    """Process all videos in the videos folder"""
    video_paths = glob('videos/*.mp4') + glob('videos/*.avi')
    video_paths.sort()
    
    if not video_paths:
        print("No videos found in the 'videos' folder!")
        return False
        
    print("\nProcessing videos...")
    success = False
    for video_path in video_paths:
        try:
            print(f"\nProcessing {video_path}")
            video_detections = process_video(video_path, sample_rate=30)  # 1 frame per second for 30fps video
            
            if video_detections:
                output_path = f'video_detections_{os.path.splitext(os.path.basename(video_path))[0]}.xlsx'
                save_to_excel(video_detections, output_path, video_mode=True)
                print(f"Results saved to {output_path}")
                success = True
            
        except Exception as e:
            print(f"Error processing {video_path}: {str(e)}")
    
    return success

def main():
    """
    Main function to process both images and videos
    """
    while True:
        choice = print_menu()
        
        if choice == '1':
            if not process_all_images():
                print("\nNo images were processed successfully.")
                
        elif choice == '2':
            if not process_all_videos():
                print("\nNo videos were processed successfully.")
                
        elif choice == '3':
            img_success = process_all_images()
            vid_success = process_all_videos()
            
            if not (img_success or vid_success):
                print("\nNo files were processed successfully.")
            
        elif choice == '4':
            print("\nExiting program. Goodbye!")
            break
            
        else:
            print("\nInvalid option! Please try again.")

if __name__ == "__main__":
    main()